# -*- coding: utf-8 -*-
import os
import sys
import csv
import glob
import fnmatch
import errno
from datetime import datetime
from openpyxl import load_workbook

import logging

today = datetime.today()
first_day = today.replace(day=1)
first_day_str = first_day.strftime('%Y-%m-%d')

#Log to recieve the data on the email
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelToCSVProccess:
    """A class to process and manage data from Excel files."""
    def __init__(self, input_folder, output_folder):
        self.input_folder = input_folder
        self.output_folder = output_folder
        self.temp_folder = os.path.join(self.input_folder, "temp_csv")
        
        # Maps desired output file names to possible sheet names from the Excel file.
        self.sheet_patterns = {
        'newlosses': ['new losses'],
        'payments': ['payment reserve activity'],
        'reserves': ['reserve changes'],
        'current_owed': ['current owed to disney'],
        'closed_claims': ['dlr closed claims', 'wdw closed claims'],
        'losses': ['wdw losses', 'dlr losses']
        }
        
        # Map only desired output files
        self.sheet_to_process = {
        'losses'
        }
        
        
        #Create a temporary folder to store and process the CSV files
        try:
            os.makedirs(self.temp_folder)
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise
        
    def get_excel_files(self,folder_path):
        
        excel_files = []
        print("The path of the folder is: {}".format(folder_path))

        #Read the excel file
        excel_files = [f for f in glob.glob(os.path.join(folder_path,"*.xlsx")) if not os.path.basename(f).startswith('~$')]
        
        for excel_file in excel_files:
            print("Excel file found: {}".format(excel_file))
            logger.info("Excel file found: {}".format(excel_file))
            self.process_excel_file(excel_file)
        
        # Consolidate CSV files
        self.consolidate_csv_files()
        
        # Clean temporary folder
        #self.cleanup_temp_files()
        
        logger.info("Process completed successfully!")
         
    def process_excel_file(self,excel_path):
        """
        Processes an Excel file by reading each sheet and saving its data as a temporary CSV.
        """
        logger.info("Processing file: {}".format(excel_path))
    
        try:
            #Creates the df for the excel file
            wb = load_workbook(filename= excel_path)
            sheet_names = wb.sheetnames
            logger.info("Sheet found: {}".format(sheet_names))
            
            base_filename = os.path.splitext(os.path.basename(excel_path))[0]
            for sheet_name in wb.sheetnames:
                
                # Identify the type of the current sheet (e.g., 'losses', 'payments') based on its name.
                sheet_type = self.identify_sheet_type(sheet_name)
                logger.info("Sheet type identified: {}".format(sheet_type))
                
                sheet_data = self.extract_sheet_data(wb[sheet_name], sheet_name, base_filename)
                
                self.save_temp_csv(sheet_data, base_filename, sheet_type)

                # df_clean.to_csv(temp_filepath,index=False,header=True)  
                # logger.info("Temp CSV saved: {}".format(temp_filename))
            
            wb.close()
            
        except Exception as e:
            logger.error("Error processing file {}: {}".format(excel_path, str(e)))

    def identify_sheet_type(self, sheet_name):
        
        # Sheet name to lower case
        sheet_lower = sheet_name.lower().strip()
        
        # Iterate through the dictionary of sheet patterns to find a matching type.
        #sheet_type = 'losses', patterns = 'wdw losses', 'dlr losses'
        for sheet_type, patterns in self.sheet_patterns.items():
            if any(pattern in sheet_lower for pattern in patterns):
                return sheet_type
        return 'unk'    
 
    def extract_sheet_data(self, worksheet, sheet_name, file_name):

        sheet_data = []
        
        try:
            #This will obtain all the data from the row
            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                row_data = []
                
                for cell_value in row:
                    formatted_value = self.get_cell_value(cell_value)
                    row_data.append(formatted_value)
                
                if row_idx == 1:
                    row_data.extend(['source_sheet', 'source_file'])
                else:
                    row_data.extend([sheet_name, file_name])
                
                sheet_data.append(row_data)
            
            logger.info("Extracted {} rows from the sheet '{}'".format(len(sheet_data), sheet_name))
            
        except Exception as e:
            logger.error("Error extracting the data from the sheet '{}': {}".format(sheet_name, str(e)))
        
        return sheet_data
    
    def get_cell_value(self, cell_value):
        if cell_value is None:
            return ''
        elif isinstance(cell_value, (int, float)):
            if isinstance(cell_value, float) and cell_value.is_integer():
                return int(cell_value)
            return cell_value
        elif isinstance(cell_value, datetime):
            return cell_value.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(cell_value, bool):
            return cell_value
        else:
            return str(cell_value)
  
    def save_temp_csv(self, sheet_data, file_prefix, sheet_type):
    
        temp_filename = f"{file_prefix}_{sheet_type}.csv"
        temp_filepath = os.path.join(self.temp_folder, temp_filename)

        try:
            # Open in text mode with a specific encoding
            with open(temp_filepath, 'w', encoding='utf-8', newline='') as csvfile:
                writer = csv.writer(csvfile)
                
                for row in sheet_data:
                    # No need to check for 'unicode' or encode manually
                    writer.writerow(row)
                    
        except Exception as e:
            logger.error(f"Error saving file {temp_filename}: {e}")
  
    def consolidate_csv_files(self) :
         
        try:
            for sheet_file in self.sheet_to_process:
            
                logger.info("Concatenating the files for : {}".format(sheet_file))
                
                concatenate_file = glob.glob(os.path.join(self.temp_folder,f"*_{sheet_file}.csv"))
                logger.info("Concatenating the files for : {}".format(concatenate_file))
                
                #Creates the final file 
                output_files =  os.path.join(self.output_folder,f"concatenate_{sheet_file}.{first_day_str}")
                total_records = 0
                
                #Open final file for writing
                with open(output_files, 'w', encoding='utf-8', newline='') as final_file:
                    writer = csv.writer(final_file)
                    headers_written = False
                    
                    #Iterate over the files found with the same pattern
                    for file in concatenate_file:
                        logger.info("Concatenating the files for : {}".format(file))
                        
                        #Open individual file to read it
                        with open(file, 'r') as temp_file:
                             reader = csv.reader(temp_file)
                             
                             for row_idx, row in enumerate(reader):
                                 if row_idx == 0 and not headers_written:
                                    writer.writerow(row)
                                    headers_written = True
                                 elif row_idx == 0 and headers_written:
                                       continue
                                 else:
                                   if row: 
                                       writer.writerow(row)
                                       total_records += 1
                
        except Exception as e:
            logger.error("Error Concatenating the files: {}".format(str(e)))     
                           
    def cleanup_temp_files(self):
        logger.info("Clean temporary files...")
        
        try:
            for filename in os.listdir(self.temp_folder):
                file_path = os.path.join(self.temp_folder, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
            
            os.rmdir(self.temp_folder)
            logger.info("Temporary files deleted")
            
        except Exception as e:
            logger.error("Error deleted temporary files: {}".format(str(e)))             

def main():
              
    input_folder = "Folder here"
    output_folder = "Folder here"
                      
    process = ExcelToCSVProccess(input_folder, output_folder)
    process.get_excel_files(input_folder)
if __name__ == "__main__":
    main()
